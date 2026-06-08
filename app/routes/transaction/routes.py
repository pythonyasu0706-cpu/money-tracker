# app/routes/transaction/routes.py
from flask import Blueprint, render_template, request, jsonify, redirect, url_for, current_app, Response
from flask_login import login_required, current_user
import cloudinary.uploader
import traceback
from datetime import datetime, timedelta, timezone
from dateutil import parser
from sqlalchemy import desc
from collections import defaultdict
from app.extensions import db
from app.models.transaction import Transaction
from app.utils.color_map import get_color_map
import io
import calendar
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

transaction_bp = Blueprint("transaction", __name__)

@transaction_bp.route("/")
def landing():
    if current_user.is_authenticated:
        return redirect(url_for("transaction.upload"))
    return render_template("landing.html")

@transaction_bp.route("/upload")
@login_required
def upload():
    return render_template("transaction/upload.html")

@transaction_bp.route('/process_receipt', methods=['POST'])
@login_required
def process_receipt():
    if 'image' not in request.files:
        return jsonify({"error": "No image provided"}), 400

    # ファイル取得
    image = request.files['image']

    # クラウド
    upload_result = cloudinary.uploader.upload(image)

    image_url = upload_result["secure_url"] #表示用URL
    public_id = upload_result["public_id"] #削除キー

    # text = current_app.ocr.extract_text(image_url)
    # result = current_app.ai.parse(text)

    # OCR処理
    try:
        text = current_app.ocr.extract_text(image_url)

    except Exception as e:
        print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500

    if text == "OCRエラー":
        return jsonify({"error": "OCR failed"}), 500

    result = current_app.ai.parse(text)

    # OCR結果から日付があれば使う（AI側）
    raw_date = result.get("date")

    receipt_date = None
    if raw_date:
        try:
            receipt_date = parser.parse(raw_date).date()
        except:
            receipt_date = None

    # 新規作成(OCRボタンを押すとここで初めてDB保存)
    expense = Transaction(
        user_id=current_user.id,
        amount=result.get("total_amount"),
        category="",
        shop_name=result.get("store_name"),
        raw_text=text,
        receipt_date=receipt_date,
        image_path=image_url, #0522変更
        image_public_id=public_id, #0522変更
        status="draft",
        type="expense",  # ★レシートは固定
        source="ocr",  # ★追加
        expires_at=datetime.now(timezone.utc) + timedelta(days=1)
    )

    db.session.add(expense)
    db.session.commit()

    return jsonify({
        "status": "ok",
        "id": expense.id
    })

# ================
# 経費登録（手入力）
# ================
@transaction_bp.route("/create_expense", methods=["POST"])
@login_required
def create_expense():
    data = request.json

    # 日付
    receipt_date = None
    if data.get("date"):
        try:
            receipt_date = parser.parse(data.get("date")).date()
        except:
            receipt_date = None

    # 金額
    amount_raw = data.get("amount", "")

    if amount_raw == "":
        return jsonify({"error": "金額を入力してください"}), 400

    try:
        amount = int(amount_raw)
    except ValueError:
        return jsonify({"error": "金額は数字で入力してください"}), 400

    # 保存
    expense = Transaction(
        user_id=current_user.id,
        amount=amount,
        category=data.get("category"),
        shop_name=data.get("store_name"),
        raw_text=data.get("ocr_text", ""),  # 手入力なら空でOK
        receipt_date=receipt_date,
        image_path=data.get("image_path"),  # 手入力はNoneでOK
        image_public_id=None, #0522変更
        status="confirmed",
        type=data.get("type", "expense"),
        source="manual",  # ★追加
        expires_at=datetime.now(timezone.utc) + timedelta(days=1)
    )

    db.session.add(expense)
    db.session.commit()

    return jsonify({
        "status": "ok",
        "id": expense.id
    })


# ================
# OCR後編集
# ================
@transaction_bp.route("/update_expense", methods=["POST"])
@login_required
def update_expense():
    data = request.json

    expense_id = data.get("id")
    if not expense_id:
        return jsonify({"error": "missing id"}), 400
            # ★更新
    expense = Transaction.query.filter_by(
        id=expense_id,
        user_id=current_user.id
    ).first_or_404()
    
    receipt_date = None
    if data.get("date"):
        try:
            receipt_date = parser.parse(data.get("date")).date()
        except:
            receipt_date = None

    print(data)  # ← とりあえず確認


    expense.amount = data.get("amount")
    expense.category = data.get("category")
    expense.shop_name = data.get("store_name")
    expense.receipt_date = receipt_date
    expense.status = "confirmed"


    db.session.commit()
    return jsonify({"status": "ok"})

# TODO: 認証機能実装後、ユーザーごとの履歴表示にする
@transaction_bp.route("/expenses", methods=["GET"])
@login_required
def get_expenses():
    expenses = Transaction.query.filter_by(user_id=current_user.id)\
        .order_by(Transaction.created_at.desc()).all()

    result = []
    for e in expenses:
        result.append({
            "id": e.id,
            "amount": e.amount,
            "category": e.category,
            "shop_name": e.shop_name,
            "raw_text": e.raw_text,
            "created_at": e.created_at.strftime("%Y-%m-%d %H:%M:%S")
        })

    return jsonify(result)

# TODO: 認証機能実装後、ユーザーごとの履歴表示にする
@transaction_bp.route("/history")
@login_required
def history():
    expenses = Transaction.query.filter_by(
        status="confirmed",
        user_id=current_user.id
        ).order_by(
            desc(Transaction.receipt_date),
            desc(Transaction.created_at)
        ).all()
    grouped = defaultdict(list)

    for e in expenses:
        date = e.receipt_date or e.created_at

        if date:
            key = date.strftime("%Y年%m月")
        else:
            key = "日付不明"

        grouped[key].append(e)
    
    # ★ここ追加（これだけ）
    grouped = dict(
        sorted(grouped.items(), key=lambda x: x[0], reverse=True)
    )

    return render_template("transaction/history.html",
                        color_map=get_color_map(),
                        grouped_expenses=grouped)

@transaction_bp.route("/debug")
def debug():
    return "THIS IS CORRECT APP"

# 履歴から編集画面へ
@transaction_bp.route("/edit/<int:id>")
@login_required
def edit_expense(id):
    expense = Transaction.query.filter_by(
        id=id,
        user_id=current_user.id
    ).first_or_404()

    # 共通データ
    data = {
        "id": expense.id,
        "status": expense.status,
        "category": expense.category,
        "type": expense.type,
        "store_name": expense.shop_name,
        "date": expense.receipt_date.strftime("%Y-%m-%d") if expense.receipt_date else "",
        "amount": expense.amount
    }

    # =====================
    # OCRルート
    # =====================
    if expense.source == "ocr":

        image_url = None
        if expense.image_path:
            image_url = expense.image_path

        data.update({
            "image_url": image_url,
            "ocr_text": expense.raw_text or "（OCRなし）",
            "result": {
                "store_name": expense.shop_name,
                "date": data["date"],
                "total_amount": expense.amount
            },
            "categories": current_app.ai.parse(expense.raw_text).get("categories", [])
        })

        return render_template("transaction/ocr_edit.html", data=data)

    # =====================
    # 手入力ルート
    # =====================
    else:
        return render_template("transaction/manual_edit.html", data=data)

@transaction_bp.route("/delete_expense/<int:id>", methods=["DELETE"])
@login_required
def delete_expense(id):
    expense = Transaction.query.filter_by(
        id=id,
        user_id=current_user.id
    ).first_or_404()

    # 画像も削除したい場合
    if expense.image_public_id:
        try:
            cloudinary.uploader.destroy(expense.image_public_id)
        except:
            pass

    db.session.delete(expense)
    db.session.commit()

    return jsonify({"status": "deleted"})

# =====================
# CSVエクスポート
# =====================
@transaction_bp.route("/export_csv")
@login_required
def export_csv():
    # フロントから開始月・終了月（例：「2026年01月」）を取得
    start_month_str = request.args.get('start_month', 'all')
    end_month_str = request.args.get('end_month', 'all')

    # 基本のクエリ
    query = Transaction.query.filter_by(
        user_id=current_user.id,
        status="confirmed"
    )

    # 開始月のフィルタリング（選択された月の「1日」以降）
    if start_month_str and start_month_str != 'all':
        try:
            start_date = datetime.strptime(start_month_str, "%Y年%m月").date()
            query = query.filter(Transaction.receipt_date >= start_date)
        except ValueError:
            pass
    else:
        start_month_str = "all"

    # 終了月のフィルタリング（選択された月の「最終日」以前）
    if end_month_str and end_month_str != 'all':
        try:
            parsed_end = datetime.strptime(end_month_str, "%Y年%m月").date()
            # 指定された月の最終日（28〜31日）を自動計算
            last_day = calendar.monthrange(parsed_end.year, parsed_end.month)[1]
            end_date = parsed_end.replace(day=last_day)
            query = query.filter(Transaction.receipt_date <= end_date)
        except ValueError:
            pass
    else:
        end_month_str = "all"

    # 日付の古い順（昇順）または新しい順（降順）でお好みで。ここでは履歴に合わせて降順
    expenses = query.order_by(Transaction.receipt_date.desc()).all()

    # --- Excel生成 ---
    wb = Workbook()
    ws = wb.active
    ws.title = "取引履歴"
    ws.views.sheetView[0].showGridLines = True

    # スタイル設定
    font_family = "Segoe UI" if "Segoe UI" in ["Segoe UI"] else "Meiryo UI"
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    body_font = Font(name=font_family, size=10)
    amount_font = Font(name=font_family, size=10, bold=True)
    
    # 月ごと「小計」用のスタイル（マイルドなグレー背景、通常の太字）
    subtotal_font = Font(name=font_family, size=10, bold=True, color="1A252F")
    subtotal_fill = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")

    # 全体「総合計」用のスタイル（少し大きめ、濃いめのシックなグレー背景、際立つ文字色）
    grandtotal_font = Font(name=font_family, size=11, bold=True, color="2C3E50")
    grandtotal_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

    thin_side = Side(style='thin', color='D3D3D3')
    medium_side = Side(style='medium', color='2C3E50') # 総合計を挟む太めの線
    double_side = Side(style='double', color='1A252F') # 総合計の最下部（2本線）
    
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    subtotal_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    grandtotal_border = Border(left=thin_side, right=thin_side, top=medium_side, bottom=double_side)

    # ヘッダー
    headers = ['日付', 'カテゴリ', '店名', '支出（円）', '収入（円）']
    ws.append(headers)
    for col_num in range(1, 6):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = cell_border
    ws.row_dimensions[1].height = 26

    # ----------------------------------------------------------------
    # データ書き込み
    # ----------------------------------------------------------------
    # データを月ごとにグループ化
    from collections import defaultdict
    monthly_data = defaultdict(list)
    for exp in expenses:
        month_key = exp.receipt_date.strftime("%Y年%m月") if exp.receipt_date else "日付不明"
        monthly_data[month_key].append(exp)

    # 期間全体の総合計カウンターを準備
    grand_expense_total = 0
    grand_income_total = 0

    current_row = 2
    
    # 月ごとに順次書き込み
    for month_title, items in sorted(monthly_data.items(), reverse=True):
        
        month_expense_total = 0
        month_income_total = 0

        for exp in items:
            date_val = exp.receipt_date.strftime("%Y/%m/%d") if exp.receipt_date else ""
            category_val = exp.category or "未分類"
            shop_val = exp.shop_name or ""
            
            expense_val = exp.amount if exp.type == "expense" else ""
            income_val = exp.amount if exp.type != "expense" else ""

            # 月ごとの電卓
            if exp.type == "expense":
                month_expense_total += exp.amount
                grand_expense_total += exp.amount # 総合計にも合算
            else:
                month_income_total += exp.amount
                grand_income_total += exp.amount # 総合計にも合算

            ws.append([date_val, category_val, shop_val, expense_val, income_val])

            ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=current_row, column=2).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row=current_row, column=3).alignment = Alignment(horizontal='left', vertical='center')
            
            for col_num in range(1, 6):
                c = ws.cell(row=current_row, column=col_num)
                c.border = cell_border
                c.font = body_font
                if col_num in [4, 5]:
                    c.alignment = Alignment(horizontal='right', vertical='center')
                    if c.value != "":
                        c.number_format = '#,##0'
                        
            ws.row_dimensions[current_row].height = 20
            current_row += 1

        # 月のデータが終わったら「小計行」を挿入
        ws.cell(row=current_row, column=3).value = f"{month_title} 小計"
        ws.cell(row=current_row, column=3).alignment = Alignment(horizontal='right', vertical='center')
        
        ws.cell(row=current_row, column=4).value = month_expense_total if month_expense_total > 0 else ""
        ws.cell(row=current_row, column=5).value = month_income_total if month_income_total > 0 else ""
        
        for col_num in range(1, 6):
            c = ws.cell(row=current_row, column=col_num)
            c.font = subtotal_font
            c.fill = subtotal_fill
            c.border = subtotal_border
            if col_num in [4, 5]:
                c.alignment = Alignment(horizontal='right', vertical='center')
                if c.value != "":
                    c.number_format = '#,##0'
                    
        ws.row_dimensions[current_row].height = 24
        current_row += 1

        # 月のブロックの間に空行を挟む
        ws.append(["", "", "", "", ""])
        current_row += 1

    # ----------------------------------------------------------------
    #【新機能】すべての月ループが終わった最下部に「総合計行」をドカンと挿入
    # ----------------------------------------------------------------
    if len(expenses) > 0:
        # 空行をスキップして、最終行を上書き・整形する形にします
        ws.cell(row=current_row-1, column=3).value = "合計" # 項目はシンプルに「合計」
        ws.cell(row=current_row-1, column=3).alignment = Alignment(horizontal='right', vertical='center')
        
        # 期間全体の合計値をセット
        ws.cell(row=current_row-1, column=4).value = grand_expense_total if grand_expense_total > 0 else ""
        ws.cell(row=current_row-1, column=5).value = grand_income_total if grand_income_total > 0 else ""
        
        # 総合計専用の特別な装飾（濃いめの背景、太い境界線）を適用
        for col_num in range(1, 6):
            c = ws.cell(row=current_row-1, column=col_num)
            c.font = grandtotal_font
            c.fill = grandtotal_fill
            c.border = grandtotal_border
            if col_num in [4, 5]:
                c.alignment = Alignment(horizontal='right', vertical='center')
                if c.value != "":
                    c.number_format = '#,##0'
                    
        ws.row_dimensions[current_row-1].height = 28

    # 列幅自動調整
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                byte_len = len(str(cell.value).encode('utf-8'))
                if byte_len > max_len:
                    max_len = byte_len
        ws.column_dimensions[col_letter].width = max(max_len // 2 + 5, 12)

    # 出力設定
    excel_out = io.BytesIO()
    wb.save(excel_out)
    excel_out.seek(0)

    response = Response(
        excel_out.getvalue(), 
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
    # ファイル名を範囲が伝わる形に綺麗に整形
    s_name = start_month_str.replace('年', '').replace('月', '')
    e_name = end_month_str.replace('年', '').replace('月', '')
    filename = f"expenses_{s_name}_to_{e_name}.xlsx"
    
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response